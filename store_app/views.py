from django.shortcuts import render,redirect
from .models import Product,Category, Profile,Charge,FormulaireCharge, FormulaireArticle,About, FormulaireClient, FormulaireVente
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm
from .forms import SignUpForm,UpdateUserForm,ChangePasswordForm, UserInfoForm, EnregistrerChargeForm,EnregistrerFormulaireChargeForm,EnregistrerFormulaireArticleForm,EnregistrerCategoryForm,AboutForm, EnregistrerFormulaireClientForm, EnregistrerFormulaireVenteForm
from django import forms
from django.db.models import Q
import json
from cart_app.cart import Cart
from payment_app.forms import ShippingForm
from payment_app.models import ShippingAddress
from django.shortcuts import get_object_or_404, redirect
from django.core.mail import send_mail,EmailMessage
from django.conf import settings
from django.core.paginator import Paginator
from django.db.models import Sum
from django.http import JsonResponse
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter
import io
from django.http import FileResponse
from django.urls import reverse
import os
import tempfile
import csv
from django.http import HttpResponse
#from excel_response import ExcelResponse  # Importa la respuesta Excel
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from datetime import date
from openpyxl.styles import Font
from django.core import management
from calendar import HTMLCalendar
from datetime import datetime
import mimetypes
from django.core.exceptions import ObjectDoesNotExist
from django.db import connection


# Create your views here.

def search(request):
    #determine if they filled out the form 
    
    if request.method == 'POST':
        searched = request.POST['searched']
        #query the products db model
        
        searched = Product.objects.filter(Q(name__icontains=searched) | Q(description__icontains=searched))
        
        #test for null
        
        if not searched:
            messages.success(request, "That Product does not xist please try again....")
            return render(request, 'store_app/search.html',{})
        else:
            context={'searched':searched}
            return render(request, 'store_app/search.html',context)
    else:
        return render(request, 'store_app/search.html',{})
    
    
    
    

def update_info(request):
    if request.user.is_authenticated:
        #get current user
        current_user= Profile.objects.get(user__id=request.user.id)
        #get current user shipping info
        shipping_user = ShippingAddress.objects.get(user__id=request.user.id)
        
        #get original user form
        form = UserInfoForm(request.POST or None, instance=current_user)
        
        #get user orignal shipping form
        shipping_form = ShippingForm(request.POST or None, instance=shipping_user)
        
        
        if form.is_valid() or shipping_form.is_valid():
            form.save()
            shipping_form.save()
            messages.success(request, "Your Info has been Updated....")
            return redirect('home')
        
            
        context={'form':form, 'shipping_form':shipping_form}
        
        return render(request, 'store_app/update_info.html',context)
    else:
        messages.success(request, "You must be logged in to access this page.......")
        return redirect('home')

def update_password(request):
    
    if request.user.is_authenticated:
        current_user=request.user
        #did the fill out the form
        if request.method == 'POST':
            foo = current_user
            form = ChangePasswordForm(current_user, request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Your password has been update please log in again Thank you....")
                login(request, current_user)
                return redirect('update_user')
            else:
                for error in list(form.errors.values()):
                    messages.error(request, error)
                return redirect('update_password')    
                    
        
        else:
            form = ChangePasswordForm(current_user)
            context={'form':form}
            return render(request, 'store_app/update_password.html',context)
    else:
        messages.success(request, "You must be logged in to view that page....")
        return redirect('home')

def update_user(request):
    
    if request.user.is_authenticated:
        current_user= User.objects.get(id=request.user.id)
        user_form = UpdateUserForm(request.POST or None, instance=current_user)
        if user_form.is_valid():
            user_form.save()
            login(request, current_user)
            messages.success(request, "User has been Updated....")
            return redirect('home')
        
            
        context={'user_form':user_form}
        
        return render(request, 'store_app/update_user.html',context)
    else:
        messages.success(request, "You must be logged in to access this page.......")
        return redirect('home')

def category_summary(request):
    categories = Category.objects.all()
    context={'categories':categories}
    return render(request, 'store_app/category_summary.html',context)

def category_detail(request, slug):
    category = get_object_or_404(Category, slug=slug)
    products = Product.objects.filter(category=category)
    context = {
        'category': category,
        'products': products,
    }
    return render(request, 'store_app/category_detail.html', context)

def category(request, foo):
    foo = foo.replace('-', ' ')
    #grab the category from url
    try:
        category=Category.objects.get(name=foo)
        products=Product.objects.filter(category=category)
        context={'products':products, 'category': category}
        return render(request, 'store_app/category.html',context)
    
    except:
        messages.success(request, ("That Cagegory Doesn't exist Sorry......"))
        return redirect('home')
    
    
    
    

def product(request, pk):
    product= Product.objects.get(id=pk)
    
    context={'product':product}
    #context = {'product': product, 'quantite_product': product.quantite_product}
    # quantity_range = range(1, product.quantite_product + 1)
    # context = {'product': product, 'quantity_range': quantity_range}

    return render(request, 'store_app/product.html',context)
    

def home(request):
    products= Product.objects.all()
    # for product in products:
    #     article = FormulaireArticle.objects.get(nom=product.name)
    #     print("Product name: " + product.name)
    
    # matching_products = Product.objects.filter(name__in=FormulaireArticle.objects.values('nom'))

    # # Iterar sobre los productos encontrados y mostrar sus nombres
    # for product in matching_products:
    #     print("Product name: " + product.name)
        
    context={'products':products}
    return render(request, 'store_app/home.html',context)

def about(request):
    context={}
    return render(request, 'store_app/about.html',context)

def create_or_update_about(request, pk=None):
    if pk:
        about = get_object_or_404(About, pk=pk)
    else:
        about = None

    if request.method == 'POST':
        form = AboutForm(request.POST, instance=about)
        if form.is_valid():
            form.save()
            return redirect('home')  # Redirige a una vista que muestre la lista de About, por ejemplo
    else:
        form = AboutForm()

    return render(request, 'about_form.html', {'form': form})

def liste_about(request):
    # venue_list = Venue.objects.all().order_by('?')
    la_lista_about = About.objects.all()
    
    name = request.user.username
    # set pagination
    
    context={'la_lista_about': la_lista_about, 'name':name}
    
    return render(request, 'store_app/about.html', context)

def actualiser_about(request, id_about):
    about = About.objects.get(pk=id_about)
    form = AboutForm(request.POST or None, request.FILES or None,  instance=about)
    if form.is_valid():
        form.save()
        return redirect('home')
    context = {'about': about, 'form': form}
    return render(request, 'store_app/actualizer_about.html', context)

def eliminer_about(request, id_about):
    about = get_object_or_404(About, id=id_about)
    about.delete()
    messages.success(request, "Le formulaire about a été eliminer correctement.")
    return redirect('liste_about')  # Reemplaza 'nombre_de_tu_vista' con el nombre de tu vista principal

def login_user(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            
            # do some shopping cart stuff
            
            current_user = Profile.objects.get(user__id=request.user.id)
            #get ther saved cart from DB
            saved_cart = current_user.old_cart
            #convert db string to python diccionary
            
            if saved_cart:
                #convert to diccionary usin json 
                converted_cart=json.loads(saved_cart)
                #add loaded dic to sesssion
                cart = Cart(request)
                
                #loop throw the cart and items form db
                
                for key, value in converted_cart.items():
                    cart.db_add(product=key, quantity=value)
                    
            
            
            messages.success(request, ("You have been logged In"))
            return redirect('home')
        else:
            messages.success(request, ("There was an error, please try again later Thanks....."))
            return redirect('login')
    else:
        context={}
        return render(request, 'store_app/login.html',context)

def logout_user(request):
    logout(request)
    messages.success(request, ("You have been logged out ... Thanks for stopping by"))
    return redirect('home')

def register_user(request):
    form = SignUpForm()
    if request.method == "POST":
        form = SignUpForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data['username']
            password = form.cleaned_data['password1']
            # log in user
            user = authenticate(username=username, password=password)
            login(request,user)
            messages.success(request, ("User Created - Please Fill out Your user infoo Below thank you....."))
            return redirect('update_info')
        
        else:
            messages.success(request, ("Whoops there was a problem registering try again please....."))
            
            return redirect('register')
    else:
        context={'form':form}
        return render(request, 'store_app/register.html',context)

def reset_table_view(request):
    table_name = 'store_app_formulairearticle'
    table_name_product = 'store_app_product'
    reset_table_and_id_counter(table_name)
    reset_table_and_id_counter(table_name_product)
    return HttpResponse(f'Tabla {table_name} reiniciada con éxito. y {table_name_product} tambien')

def reset_table_and_id_counter(table_name):
    with connection.cursor() as cursor:
        # Vaciar la tabla
        cursor.execute(f"DELETE FROM {table_name};")
        # Reiniciar el contador de ID
        cursor.execute(f"DELETE FROM sqlite_sequence WHERE name='{table_name}';")  

def enregistrer_charge_view(request):
    if request.method == 'POST':
        form = EnregistrerChargeForm(request.POST, request.FILES)
        if form.is_valid():
            #form.save()
            #titulo_serie = request.POST['titulo_serie']
            #serie_o_pelicula = request.POST['serie_o_pelicula']
            #plataforma = request.POST['plataforma']
            #name = request.user.username
            #file = request.FILES['file']
            instance = form.save()
            nome_charge = instance.nome_charge
           
            #files = request.FILES.getlist('files')
            email_message = EmailMessage(
                subject=f'Contact Form: {nome_charge} - {nome_charge}',
                #body=titulo_serie + " " + serie_o_pelicula + " " +  plataforma,
                body=f'Nom de la charge: {nome_charge}\nNom de la charge: {nome_charge}\nNom de la charge: {nome_charge}',
                
                from_email=settings.EMAIL_HOST_USER,
                to=['msb.duck@gmail.com', 'msb.tesla@gmail.com', 'msebti2@gmail.com', 'msb.acer@gmail.com'],
                reply_to=['msebti2@gmail.com']
            )
            # Adjuntar cada archivo
            #for file in files:
                #email_message.attach(file.name, file.read(), file.content_type)
            
            #if serie_pelicula_imagen:
                #mime_type, _ = mimetypes.guess_type(serie_pelicula_imagen.path)
                #email_message.attach(serie_pelicula_imagen.name, serie_pelicula_imagen.read(), mime_type)
            
            # Adjuntar el archivo
            #email_message.attach(file.name, file.read(), file.content_type)

            # Enviar el email
            email_message.send(fail_silently=False)
            form.save()
            return redirect('liste_des_charges')  # Cambia esto por la vista a la que deseas redirigir después de guardar
    else:
        form = EnregistrerChargeForm()
    return render(request, 'enregistrer_charge.html', {'form': form})

def liste_des_charges(request):
    # venue_list = Venue.objects.all().order_by('?')
    la_lista_des_charges = Charge.objects.all()
    name = request.user.username
    # set pagination
    
    p = Paginator(la_lista_des_charges, 5)
    page = request.GET.get('page')
    tous_les_charges = p.get_page(page)
    nums = "a" * tous_les_charges.paginator.num_pages
    
    print("hola : " + str(tous_les_charges.paginator.num_pages))
    
    context = {'la_lista_des_charges': la_lista_des_charges, 'tous_les_charges': tous_les_charges, 'nums': nums, 'name':name}
    return render(request, 'store_app/la_liste_des_charges.html', context)

def actualiser_la_charge(request, id_charge):
    charge = Charge.objects.get(pk=id_charge)
    form = EnregistrerChargeForm(request.POST or None, request.FILES or None,  instance=charge)
    if form.is_valid():
        form.save()
        messages.success(request, "La charge a été actualisée correctement.")
        return redirect('liste_des_charges')
    context = {'charge': charge, 'form': form}
    return render(request, 'store_app/actualizer_la_charge.html', context)

def eliminer_la_charge(request, id_charge):
    charge = get_object_or_404(Charge, id=id_charge)
    charge.delete()
    messages.success(request, "La charge a été eliminer correctement.")
    return redirect('liste_des_charges')  # Reemplaza 'nombre_de_tu_vista' con el nombre de tu vista principal

def liste_des_formulaire_charges(request):
    # venue_list = Venue.objects.all().order_by('?')
    la_lista_des_formulaire_charges = FormulaireCharge.objects.all()
    total_montant = la_lista_des_formulaire_charges.aggregate(total=Sum('prix'))['total']
    name = request.user.username
    # set pagination
    
    p = Paginator(la_lista_des_formulaire_charges, 5)
    page = request.GET.get('page')
    tous_les_formulaire_charges = p.get_page(page)
    nums = "a" * tous_les_formulaire_charges.paginator.num_pages
    
    print("hola : " + str(tous_les_formulaire_charges.paginator.num_pages))
    
    context = {'la_lista_des_formulaire_charges': la_lista_des_formulaire_charges, 'tous_les_formulaire_charges': tous_les_formulaire_charges, 'nums': nums, 'name':name, 'total_montant':total_montant}
    
    if request.GET.get('export') == 'excel':
        # Crear un libro de trabajo y una hoja
        today = datetime.today()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Formulaire Concierge"
        
        # Crear una nueva hoja para la tabla
        #ws2 = wb.create_sheet(title="Formulaire Concierge")

        # Agregar los encabezados de la tabla
        headers = ['Date', 'Charge', 'Prix', 'Nom Fichier']
        #if request.user.is_superuser:
            #headers.extend(['Imprimer pdf', 'Actualiser', 'Eliminer'])
        ws.append(headers)

        # Aplicar negrita a los encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Agregar los datos de la tabla
        for item in la_lista_des_formulaire_charges:
            row = [
                item.date.strftime('%Y-%m-%d'),
                item.charge.nome_charge,
                item.prix,
                item.image_charge.name
            ]
            #if request.user.is_superuser:
                #row.extend(['Imprimer', 'Actualiser', 'Eliminer'])
            ws.append(row)
            
        ws.append(['Total Montant', '', total_montant])
        
        # Crear una respuesta HTTP con el archivo Excel
        filename = f"liste_charges_immeuble_{today.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        #response['Content-Disposition'] = 'attachment; filename=lista_paie_concierge.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
    return render(request, 'store_app/la_liste_des_formulaire_charges.html', context)

def enregistrer_formulaire_charge_view(request):
    if request.method == 'POST':
        form = EnregistrerFormulaireChargeForm(request.POST, request.FILES)
        if form.is_valid():
            #form.save()
            #titulo_serie = request.POST['titulo_serie']
            #serie_o_pelicula = request.POST['serie_o_pelicula']
            #plataforma = request.POST['plataforma']
            #name = request.user.username
            #file = request.FILES['file']
            instance = form.save()
            charge = instance.charge
           
            date = instance.date
            prix=instance.prix
            image_charge = instance.image_charge
           
            #files = request.FILES.getlist('files')
            email_message = EmailMessage(
                subject=f'Contact Form: {date} - {charge} - {prix}',
                #body=titulo_serie + " " + serie_o_pelicula + " " +  plataforma,
                body=f'Nom de la charge: {charge}\nDate de la charge: {date}\nPrix de la charge: {prix}',
                
                from_email=settings.EMAIL_HOST_USER,
                to=['msb.duck@gmail.com', 'msb.tesla@gmail.com', 'msebti2@gmail.com', 'msb.acer@gmail.com'],
                reply_to=['msebti2@gmail.com']
            )
            # Adjuntar cada archivo
            #for file in files:
                #email_message.attach(file.name, file.read(), file.content_type)
            
            if image_charge:
                mime_type, _ = mimetypes.guess_type(image_charge.path)
                email_message.attach(image_charge.name, image_charge.read(), mime_type)
            
            # Adjuntar el archivo
            #email_message.attach(file.name, file.read(), file.content_type)

            # Enviar el email
            email_message.send(fail_silently=False)
            form.save()
            return redirect('liste_des_formulaire_charges')  # Cambia esto por la vista a la que deseas redirigir después de guardar
    else:
        form = EnregistrerFormulaireChargeForm()
    return render(request, 'enregistrer_formulaire_charge.html', {'form': form})

def actualiser_formulaire_charge(request, id_formulaire_charge):
    formulaire_charge = FormulaireCharge.objects.get(pk=id_formulaire_charge)
    form = EnregistrerFormulaireChargeForm(request.POST or None, request.FILES or None,  instance=formulaire_charge)
    if form.is_valid():
        form.save()
        return redirect('liste_des_formulaire_charges')
    context = {'formulaire_charge': formulaire_charge, 'form': form}
    return render(request, 'store_app/actualizer_formulaire_charge.html', context)

def eliminer_formulaire_charge(request, id_formulaire_charge):
    formulaire_charge = get_object_or_404(FormulaireCharge, id=id_formulaire_charge)
    formulaire_charge.delete()
    messages.success(request, "Le formulaire charge a été eliminer correctement.")
    return redirect('liste_des_formulaire_charges')  # Reemplaza 'nombre_de_tu_vista' con el nombre de tu vista principal


def liste_des_formulaire_articles(request):
    # venue_list = Venue.objects.all().order_by('?')
    la_lista_des_formulaire_articles = FormulaireArticle.objects.all()
    total_montant = la_lista_des_formulaire_articles.aggregate(total=Sum('prix'))['total']
    name = request.user.username
    # set pagination
    
    p = Paginator(la_lista_des_formulaire_articles, 5)
    page = request.GET.get('page')
    tous_les_formulaire_articles = p.get_page(page)
    nums = "a" * tous_les_formulaire_articles.paginator.num_pages
    
    print("hola : " + str(tous_les_formulaire_articles.paginator.num_pages))
    
    context = {'la_lista_des_formulaire_articles': la_lista_des_formulaire_articles, 'tous_les_formulaire_articles': tous_les_formulaire_articles, 'nums': nums, 'name':name, 'total_montant':total_montant}
    
    if request.GET.get('export') == 'excel':
        # Crear un libro de trabajo y una hoja
        today = datetime.today()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Formulaire Article"
        
        # Crear una nueva hoja para la tabla
        #ws2 = wb.create_sheet(title="Formulaire Concierge")

        # Agregar los encabezados de la tabla
        headers = ['Ref Article', 'Quantité Article','Nom Article', 'Description Article', 'Prix','Cree le','Vendu', 'Nom Fichier']
        #if request.user.is_superuser:
            #headers.extend(['Imprimer pdf', 'Actualiser', 'Eliminer'])
        ws.append(headers)

        # Aplicar negrita a los encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Agregar los datos de la tabla
        for item in la_lista_des_formulaire_articles:
            row = [
                item.ref_article,
                item.quantite_article,
                item.nom,
                
                item.description,
                item.prix,
                item.cree_le.strftime('%Y-%m-%d'),
                item.vendu,
                item.category.name,
                item.image_charge.name
            ]
            #if request.user.is_superuser:
                #row.extend(['Imprimer', 'Actualiser', 'Eliminer'])
            ws.append(row)
            
        ws.append(['Total Montant', '',  total_montant])
        
        # Crear una respuesta HTTP con el archivo Excel
        filename = f"liste_articles_{today.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        #response['Content-Disposition'] = 'attachment; filename=lista_paie_concierge.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
    return render(request, 'store_app/la_lista_des_formulaire_articles.html', context)

def enregistrer_formulaire_article_view(request, pk=None):
    if pk:
        article = get_object_or_404(FormulaireArticle, pk=pk)
        form = EnregistrerFormulaireArticleForm(instance=article)
    else:
        # Obtener el próximo ID disponible
        #next_id = FormulaireArticle.objects.count() + 1
        #ref_article_value = f"OZ{next_id:03d}"
        
        form = EnregistrerFormulaireArticleForm()
    
    if request.method == 'POST':
        if pk:
            # Editar un artículo existente
            article = get_object_or_404(FormulaireArticle, pk=pk)
            form = EnregistrerFormulaireArticleForm(request.POST, request.FILES, instance=article)
        else:
            # Crear un nuevo artículo
            form = EnregistrerFormulaireArticleForm(request.POST, request.FILES)
            # Establecer el valor de ref_article después de validar el formulario
            #form.initial = {'ref_article': ref_article_value}
        
        if form.is_valid():
            try:
            
                #form.save()
                #titulo_serie = request.POST['titulo_serie']
                #serie_o_pelicula = request.POST['serie_o_pelicula']
                #plataforma = request.POST['plataforma']
                #name = request.user.username
                #file = request.FILES['file']
                # Verificar que la categoría existe
                
                
                category_id = request.POST.get('category')
                print("le categori_id es :" + category_id)
                
                if category_id:
                    try:
                        category = Category.objects.get(id=category_id)
                        print(category.name)
                        print(category)
                    except ObjectDoesNotExist:
                        return HttpResponse(f"Error: La categoría con id {category_id} no existe.")
                
                instance = form.save(commit=False)
                instance.category_id = request.POST.get('category')
                instance.save()
                category = instance.category
                print("esta categoria :" + category.name)
                ref_article=instance.ref_article
                nom = instance.nom
                description=instance.description
                cout_revient_ozaz=instance.cout_revient_ozaz
                cout_revient_maallem=instance.cout_revient_maallem
                cout_revient_total=instance.cout_revient_total
                
                prix=instance.prix
                cree_le = instance.cree_le
                vendu=instance.vendu
                image_charge = instance.image_charge
            
                #files = request.FILES.getlist('files')
                email_message = EmailMessage(
                    subject=f'Contact Form: {date} - {nom} - {prix}',
                    #body=titulo_serie + " " + serie_o_pelicula + " " +  plataforma,
                    body=f'Reference du Article: {ref_article}\nNom du Article: {nom}\nDescription du article: {description}\nPrix de article: {prix}\nDate creation article: {cree_le}\nEtat du article: {vendu}\nCategory: {category.name}',
                    
                    from_email=settings.EMAIL_HOST_USER,
                    to=['msb.duck@gmail.com', 'msb.tesla@gmail.com', 'msebti2@gmail.com', 'msb.acer@gmail.com'],
                    reply_to=['msebti2@gmail.com']
                )
                # Adjuntar cada archivo
                #for file in files:
                    #email_message.attach(file.name, file.read(), file.content_type)
                
                if image_charge:
                    mime_type, _ = mimetypes.guess_type(image_charge.path)
                    email_message.attach(image_charge.name, image_charge.read(), mime_type)
                
                # Adjuntar el archivo
                #email_message.attach(file.name, file.read(), file.content_type)

                # Enviar el email
                email_message.send(fail_silently=False)
                form.save()
                print(f"Form Data: {form.cleaned_data}")
                print(f"Saved Instance: {instance.ref_article}")
                
                return redirect('liste_des_formulaire_articles')  # Cambia esto por la vista a la que deseas redirigir después de guardar
            except Category.DoesNotExist as e:
                return HttpResponse(f"Error: {str(e)}")
        else:
            # Imprimir errores de formulario para depuración
            return HttpResponse(f"Errores en el formulario: {form.errors}")
        
    else:
        form = EnregistrerFormulaireArticleForm()
    return render(request, 'enregistrer_formulaire_article.html', {'form': form})

def actualiser_formulaire_article(request, id_formulaire_article):
    formulaire_article = FormulaireArticle.objects.get(pk=id_formulaire_article)
    form = EnregistrerFormulaireArticleForm(request.POST or None, request.FILES or None,  instance=formulaire_article)
    if form.is_valid():
        form.save()
        messages.success(request, "Le formulaire article a été actualisé correctement.")
        return redirect('liste_des_formulaire_articles')
    context = {'formulaire_article': formulaire_article, 'form': form}
    return render(request, 'store_app/actualizer_formulaire_article.html', context)

def eliminer_formulaire_article(request, id_formulaire_article):
    formulaire_article = get_object_or_404(FormulaireArticle, id=id_formulaire_article)
    formulaire_article.delete()
    messages.success(request, "Le formulaire article a été eliminer correctement.")
    return redirect('liste_des_formulaire_articles')  # Reemplaza 'nombre_de_tu_vista' con el nombre de tu vista principal


def liste_situation_caisse(request):
    # venue_list = Venue.objects.all().order_by('?')
    #la_lista_formulaire_cotization = FormulaireCotization.objects.all()
    la_lista_formulaire_article = FormulaireArticle.objects.all()
    total_montant_article = la_lista_formulaire_article.aggregate(total=Sum('prix'))['total']
    
   
     
    la_lista_des_formulaire_charges = FormulaireCharge.objects.all()
    total_montant_charge = la_lista_des_formulaire_charges.aggregate(total=Sum('prix'))['total']
    
    la_lista_des_formulaire_ventes = FormulaireVente.objects.filter(hidden_vendu_vente=True)
    #la_lista_des_formulaire_ventes = FormulaireVente.objects.all()
    total_montant_ventes = la_lista_des_formulaire_ventes.aggregate(total=Sum('prix_vente'))['total']
    
    montant_por_vente = {}
    for formulaire_vente in la_lista_des_formulaire_ventes:
        if formulaire_vente.etat_final == 'Espece':
            if 'Espece' in montant_por_vente:
                montant_por_vente['Espece'] += formulaire_vente.prix_vente
            else:
                montant_por_vente['Espece'] = formulaire_vente.prix_vente
        else:
            if 'Non_Espece' in montant_por_vente:
                montant_por_vente['Non_Espece'] += formulaire_vente.prix_vente
            else:
                montant_por_vente['Non_Espece'] = formulaire_vente.prix_vente

    # Convertir el diccionario en listas para usar en la plantilla
    montant_por_vente_list = [{'espece': key, 'prix_vente': value} for key, value in montant_por_vente.items()]

    if total_montant_ventes is None  :
        total_montant_ventes = 0

  
    
    if total_montant_article is None  :
        total_montant_article = 0
    
    if total_montant_charge is None  :
        total_montant_charge = 0
    
    #total =  total_montant_cotizacion - total_montant_charge   
    # Calculating total only if both values are not None
    if total_montant_article is not None and total_montant_charge is not None :
        total = total_montant_ventes - total_montant_charge
    else:
        total = 0  # Default to 0 if either value is None
    name = request.user.username
    
    
    
    context = {
        
        'la_lista_formulaire_article': la_lista_formulaire_article, 
        'la_lista_des_formulaire_charges': la_lista_des_formulaire_charges, 
        'name': name, 
        'total_montant_article': total_montant_article,
        'total_montant_charge': total_montant_charge, 
        'total': total, 
        'total_montant_ventes': total_montant_ventes,
        'montant_por_vente_list': montant_por_vente_list
    }    
    if request.GET.get('export') == 'excel':
        # Crear un libro de trabajo y una hoja
        today = date.today()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Situation Caisse"
        
        ws.append(['Date', today.strftime('%Y-%m-%d')])
        # Agregar los datos a la hoja
        ws.append(['Montant Charge', total_montant_charge, " dh"])
        ws.append(['Montant Article', total_montant_article , " dh"])
        ws.append(['Total ( Cotisation - Article )', total , " dh"])
        
        # Aplicar negrita a la celda "Date"
        cell = ws['B1']
        cell.font = Font(bold=True)
        
        # Aplicar negrita a toda la columna A
        for cell in ws['A']:
            cell.font = Font(bold=True)
        filename = f"situation_caisse_{today.strftime('%Y%m%d_%H%M%S')}.xlsx"
        # Crear una respuesta HTTP con el archivo Excel
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    
    return render(request, 'store_app/la_lista_situation_caisse.html', context)


def enregistrer_category_view(request):
    if request.method == 'POST':
        form = EnregistrerCategoryForm(request.POST, request.FILES)
        if form.is_valid():
            #form.save()
            #titulo_serie = request.POST['titulo_serie']
            #serie_o_pelicula = request.POST['serie_o_pelicula']
            #plataforma = request.POST['plataforma']
            #name = request.user.username
            #file = request.FILES['file']
            instance = form.save()
            name = instance.name
           
            #files = request.FILES.getlist('files')
            email_message = EmailMessage(
                subject=f'Contact Form: {name} - {name}',
                #body=titulo_serie + " " + serie_o_pelicula + " " +  plataforma,
                body=f'Nom de la category: {name}',
                
                from_email=settings.EMAIL_HOST_USER,
                to=['msb.duck@gmail.com', 'msb.tesla@gmail.com', 'msebti2@gmail.com', 'msb.acer@gmail.com'],
                reply_to=['msebti2@gmail.com']
            )
            # Adjuntar cada archivo
            #for file in files:
                #email_message.attach(file.name, file.read(), file.content_type)
            
            #if serie_pelicula_imagen:
                #mime_type, _ = mimetypes.guess_type(serie_pelicula_imagen.path)
                #email_message.attach(serie_pelicula_imagen.name, serie_pelicula_imagen.read(), mime_type)
            
            # Adjuntar el archivo
            #email_message.attach(file.name, file.read(), file.content_type)

            # Enviar el email
            email_message.send(fail_silently=False)
            form.save()
            return redirect('liste_des_categories')  # Cambia esto por la vista a la que deseas redirigir después de guardar
    else:
        form = EnregistrerCategoryForm()
    return render(request, 'enregistrer_category.html', {'form': form})

def liste_des_categories(request):
    # venue_list = Venue.objects.all().order_by('?')
    la_lista_des_categories = Category.objects.all()
    name = request.user.username
    # set pagination
    
    p = Paginator(la_lista_des_categories, 5)
    page = request.GET.get('page')
    tous_les_categories = p.get_page(page)
    nums = "a" * tous_les_categories.paginator.num_pages
    
    print("hola : " + str(tous_les_categories.paginator.num_pages))
    
    context = {'la_lista_des_categories': la_lista_des_categories, 'tous_les_categories': tous_les_categories, 'nums': nums, 'name':name}
    return render(request, 'store_app/la_lista_des_categories.html', context)

def actualiser_la_category(request, id_category):
    category = Category.objects.get(pk=id_category)
    form = EnregistrerCategoryForm(request.POST or None, request.FILES or None,  instance=category)
    if form.is_valid():
        form.save()
        messages.success(request, "La category a été actualisé correctement.")
        return redirect('liste_des_categories')
    context = {'category': category, 'form': form}
    return render(request, 'store_app/actualizer_la_category.html', context)

def eliminer_la_category(request, id_category):
    category = get_object_or_404(Category, id=id_category)
    category.delete()
    messages.success(request, "La category a été eliminer correctement.")
    return redirect('liste_des_categories')  # Reemplaza 'nombre_de_tu_vista' con el nombre de tu vista principal


def liste_des_formulaire_clients(request):
    # venue_list = Venue.objects.all().order_by('?')
    la_lista_des_formulaire_clients = FormulaireClient.objects.all()

    name = request.user.username
    # set pagination
    
    p = Paginator(la_lista_des_formulaire_clients, 5)
    page = request.GET.get('page')
    tous_les_formulaire_clients = p.get_page(page)
    nums = "a" * tous_les_formulaire_clients.paginator.num_pages
    
    print("hola : " + str(tous_les_formulaire_clients.paginator.num_pages))
    
    context = {'la_lista_des_formulaire_clients': la_lista_des_formulaire_clients, 'tous_les_formulaire_clients': tous_les_formulaire_clients, 'nums': nums, 'name':name}
    
    if request.GET.get('export') == 'excel':
        # Crear un libro de trabajo y una hoja
        today = datetime.today()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Formulaire Client"
        
        # Crear una nueva hoja para la tabla
        #ws2 = wb.create_sheet(title="Formulaire Concierge")

        # Agregar los encabezados de la tabla
        headers = ['Nom', 'Prenom', 'Tel', 'Email','Addresse', 'Description', 'Cree le']
        #if request.user.is_superuser:
            #headers.extend(['Imprimer pdf', 'Actualiser', 'Eliminer'])
        ws.append(headers)

        # Aplicar negrita a los encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Agregar los datos de la tabla
        for item in la_lista_des_formulaire_clients:
            row = [
                item.nom,
                item.prenom,
                item.tel,
                item.email,
                item.addresse,
                item.description,
                item.cree_le.strftime('%Y-%m-%d'),
               
            ]
            #if request.user.is_superuser:
                #row.extend(['Imprimer', 'Actualiser', 'Eliminer'])
            ws.append(row)
            
                
        # Crear una respuesta HTTP con el archivo Excel
        filename = f"liste_clients_{today.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        #response['Content-Disposition'] = 'attachment; filename=lista_paie_concierge.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
    return render(request, 'store_app/la_lista_des_formulaire_clients.html', context)

def enregistrer_formulaire_client_view(request):
    if request.method == 'POST':
        form = EnregistrerFormulaireClientForm(request.POST, request.FILES)
        if form.is_valid():
            try:
            
                #form.save()
                #titulo_serie = request.POST['titulo_serie']
                #serie_o_pelicula = request.POST['serie_o_pelicula']
                #plataforma = request.POST['plataforma']
                #name = request.user.username
                #file = request.FILES['file']
                # Verificar que la categoría existe

                instance = form.save(commit=False)
                instance.save()
                
                nom = instance.nom
                prenom=instance.prenom
                tel=instance.tel
                email=instance.email 
                addresse=instance.addresse
                description=instance.description
                cree_le = instance.cree_le
            
                #files = request.FILES.getlist('files')
                email_message = EmailMessage(
                    subject=f'Contact Form: {date} - {nom} - {prenom}',
                    #body=titulo_serie + " " + serie_o_pelicula + " " +  plataforma,
                    body=f'Nom : {nom}\nPrenom: {prenom}\nTel: {tel}\nEmail: {email}\nAddresse: {addresse}\nDescription: {description}\nDate creation client: {cree_le}',
                    
                    from_email=settings.EMAIL_HOST_USER,
                    to=['msb.duck@gmail.com', 'msb.tesla@gmail.com', 'msebti2@gmail.com', 'msb.acer@gmail.com'],
                    reply_to=['msebti2@gmail.com']
                )
                # Adjuntar cada archivo
                #for file in files:
                    #email_message.attach(file.name, file.read(), file.content_type)
                
                # Adjuntar el archivo
                #email_message.attach(file.name, file.read(), file.content_type)

                # Enviar el email
                email_message.send(fail_silently=False)
                form.save()
                return redirect('liste_des_formulaire_clients')  # Cambia esto por la vista a la que deseas redirigir después de guardar
            except Category.DoesNotExist as e:
                return HttpResponse(f"Error: {str(e)}")
        else:
            # Imprimir errores de formulario para depuración
            return HttpResponse(f"Errores en el formulario: {form.errors}")
        
    else:
        form = EnregistrerFormulaireClientForm()
    return render(request, 'enregistrer_formulaire_client.html', {'form': form})


def actualiser_formulaire_client(request, id_formulaire_client):
    formulaire_client = FormulaireClient.objects.get(pk=id_formulaire_client)
    form = EnregistrerFormulaireClientForm(request.POST or None, request.FILES or None,  instance=formulaire_client)
    if form.is_valid():
        form.save()
        messages.success(request, "Le formulaire client a été actualisé correctement.")
        return redirect('liste_des_formulaire_clients')
    context = {'formulaire_client': formulaire_client, 'form': form}
    return render(request, 'store_app/actualizer_formulaire_client.html', context)

def eliminer_formulaire_client(request, id_formulaire_client):
    formulaire_client = get_object_or_404(FormulaireClient, id=id_formulaire_client)
    formulaire_client.delete()
    messages.success(request, "Le formulaire client a été eliminer correctement.")
    return redirect('liste_des_formulaire_clients')  # Reemplaza 'nombre_de_tu_vista' con el nombre de tu vista principal

def remove_timezone(dt):
    """Elimina la zona horaria de un objeto datetime."""
    if dt is not None and hasattr(dt, 'tzinfo'):
        return dt.replace(tzinfo=None)
    return dt

def liste_des_formulaire_ventes(request):
    # venue_list = Venue.objects.all().order_by('?')
    la_lista_des_formulaire_ventes = FormulaireVente.objects.all()

    name = request.user.username
    # set pagination
    
    p = Paginator(la_lista_des_formulaire_ventes, 5)
    page = request.GET.get('page')
    tous_les_formulaire_ventes = p.get_page(page)
    nums = "a" * tous_les_formulaire_ventes.paginator.num_pages
    
    print("hola : " + str(tous_les_formulaire_ventes.paginator.num_pages))
    
    context = {'la_lista_des_formulaire_ventes': la_lista_des_formulaire_ventes, 'tous_les_formulaire_ventes': tous_les_formulaire_ventes, 'nums': nums, 'name':name}
    
    if request.GET.get('export') == 'excel':
        # Crear un libro de trabajo y una hoja
        today = datetime.today()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Formulaire Vente"
        
        # Crear una nueva hoja para la tabla
        #ws2 = wb.create_sheet(title="Formulaire Concierge")

        # Agregar los encabezados de la tabla
        headers = ['Article Vent Ref','Article Vente Nom', 'Date Vente', 'Client Vente nom', 'Client Vente prenom', 'Description Vente','Category Vente', 'Coût revient Total Vente', 'Prix Estimé', 'Etat Vente', 'Etat Livraison', 'Etat Payement', 'Etat Final', 'vendu_vente', 'Cree le']
        #if request.user.is_superuser:
            #headers.extend(['Imprimer pdf', 'Actualiser', 'Eliminer'])
        ws.append(headers)

        # Aplicar negrita a los encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Agregar los datos de la tabla
        for item in la_lista_des_formulaire_ventes:
            row = [
                item.article_vente.ref_article,
                item.article_vente.nom,
                remove_timezone(item.date_vente),
                item.client_vente.nom,
                item.client_vente.prenom,
                item.description_vente,
                item.category_vente,
                item.cout_revient_total_vente,
                item.prix_vente,
                item.etat_vente,
                item.etat_livrer,
                item.etat_payer,
                item.etat_final,
                item.vendu_vente,
                item.cree_le.strftime('%Y-%m-%d'),
               
            ]
            #if request.user.is_superuser:
                #row.extend(['Imprimer', 'Actualiser', 'Eliminer'])
            ws.append(row)
            
                
        # Crear una respuesta HTTP con el archivo Excel
        filename = f"liste_ventes_{today.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        #response['Content-Disposition'] = 'attachment; filename=lista_paie_concierge.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
    return render(request, 'store_app/la_lista_des_formulaire_ventes.html', context)


def enregistrer_formulaire_vente_view(request):
    if request.method == 'POST':
        form = EnregistrerFormulaireVenteForm(request.POST, request.FILES)
       
        if form.is_valid():
            try:
                print(request.POST)
                form_data = form.cleaned_data
                #print(form_data['vendu_vente']) 
                print("--------------------------")
                #print("Campos del formulario:", form.fields.keys())
                #print("Valor de vendu_vente en cleaned_data:", form.cleaned_data.get('vendu_vente'))
                quantite_vente = form_data['quantite_vente']
                #print("Datos del formulario:", form.cleaned_data)
                # Guardar la venta
                instance = form.save(commit=False)
                print("Vendu Vente:", form.cleaned_data.get('vendu_vente'))
                
                # Verificar la cantidad disponible
                #quantite_vente = instance.quantite_vente
                #print("mohammed   " + str(quantite_vente))
                article_vente = instance.article_vente.ref_article
                instance.cout_revient_total_vente *= quantite_vente
                instance.prix_vente *= quantite_vente
                instance.marge_chifre *= quantite_vente
                
                if quantite_vente > instance.article_vente.quantite_article:
                    messages.error(request, f"La quantité demandée ({quantite_vente}) dépasse la quantité disponible en stock ({article_vente}).")
                    return render(request, 'enregistrer_formulaire_vente.html', {'form': form})
                
                 # Guardar la venta
                instance.save()
                
                #print("Venta guardada:", instance)

                # Actualizar la cantidad del artículo
                #article_vente.quantite_article -= quantite_vente
                #article_vente.save()
                #print("Cantidad del artículo actualizada:", article_vente.quantite_article)
                
                          

                # Enviar el email
                email_message = EmailMessage(
                    subject=f'Contact Form: {instance.date_vente} - {instance.article_vente.ref_article} - {instance.date_vente}',
                    body=f'client_vente : {instance.client_vente}\ndescription_vente: {instance.description_vente}\ncategory_vente: {instance.category_vente}\nprix_vente: {instance.prix_vente}\nDate creation client: {instance.cree_le}',
                    from_email=settings.EMAIL_HOST_USER,
                    to=['msb.duck@gmail.com', 'msb.tesla@gmail.com', 'msebti2@gmail.com', 'msb.acer@gmail.com'],
                    reply_to=['msebti2@gmail.com']
                )
                email_message.send(fail_silently=False)

                return redirect('liste_des_formulaire_ventes')
            except Category.DoesNotExist as e:
                return HttpResponse(f"Error: {str(e)}")
            except Exception as e:
                messages.error(request, f"Une erreur s'est produite: {str(e)}")
                return render(request, 'enregistrer_formulaire_vente.html', {'form': form})
        else:
            messages.error(request, "Il y a des erreurs dans le formulaire. Veuillez corriger les erreurs et réessayer.")
            return render(request, 'enregistrer_formulaire_vente.html', {'form': form})
    else:
        form = EnregistrerFormulaireVenteForm()
        return render(request, 'enregistrer_formulaire_vente.html', {'form': form})

def enregistrer_formulaire_vente_view2(request):
    if request.method == 'POST':
        form = EnregistrerFormulaireVenteForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                quantite_vente = form.cleaned_data['quantite_vente']
                article_vente = form.cleaned_data['article_vente']
            
                #form.save()
                #titulo_serie = request.POST['titulo_serie']
                #serie_o_pelicula = request.POST['serie_o_pelicula']
                #plataforma = request.POST['plataforma']
                #name = request.user.username
                #file = request.FILES['file']
                # Verificar que la categoría existe
                if quantite_vente > article_vente.quantite_article:
                    messages.error(request, f"La quantité demandée ({quantite_vente}) dépasse la quantité disponible en stock ({article_vente.quantite_article}).")
                    return render(request, 'enregistrer_formulaire_vente.html', {'form': form})

                else:
                    # Guardar la venta
                    instance = form.save(commit=False)
                    instance.save()
                    
                    #instance = form.save(commit=False)
                    #instance.save()
                    
                    article_vente = instance.article_vente
                    quantite_vente = instance.quantite_vente
                    date_vente = instance.date_vente
                    client_vente = instance.client_vente
                    description_vente = instance.description_vente
                    category_vente = instance.category_vente
                    cout_revient_total_vente = instance.cout_revient_total_vente
                    prix_vente = instance.prix_vente
                    etat_vente = instance.etat_vente
                    etat_livrer = instance.etat_livrer
                    etat_payer = instance.etat_payer
                    etat_final = instance.etat_final
                    cree_le = instance.cree_le
                
                    #files = request.FILES.getlist('files')
                    email_message = EmailMessage(
                        subject=f'Contact Form: {date} - {article_vente} - {date_vente}',
                        #body=titulo_serie + " " + serie_o_pelicula + " " +  plataforma,
                        body=f'client_vente : {client_vente}\ndescription_vente: {description_vente}\ncategory_vente: {category_vente}\nprix_vente: {prix_vente}\nDate creation client: {cree_le}',
                        
                        from_email=settings.EMAIL_HOST_USER,
                        to=['msb.duck@gmail.com', 'msb.tesla@gmail.com', 'msebti2@gmail.com', 'msb.acer@gmail.com'],
                        reply_to=['msebti2@gmail.com']
                    )
                    # Adjuntar cada archivo
                    #for file in files:
                        #email_message.attach(file.name, file.read(), file.content_type)
                    
                    # Adjuntar el archivo
                    #email_message.attach(file.name, file.read(), file.content_type)

                    # Enviar el email
                    email_message.send(fail_silently=False)
                    form.save()
                    
                    return redirect('liste_des_formulaire_ventes')  # Cambia esto por la vista a la que deseas redirigir después de guardar
            except Category.DoesNotExist as e:
                return HttpResponse(f"Error: {str(e)}")
            """ except Exception as e:
                messages.error(request, f"Une erreur s'est produite: {str(e)}")
                return render(request, 'enregistrer_formulaire_vente.html', {'form': form}) """
        
        else:
            # Imprimir errores de formulario para depuración
            #return HttpResponse(f"Errores en el formulario: {form.errors}")
             # Si el formulario no es válido, mostrar los errores en la misma página del formulario
            messages.error(request, "Il y a des erreurs dans le formulaire. Veuillez corriger les erreurs et réessayer.")
            return render(request, 'enregistrer_formulaire_vente.html', {'form': form})
    
        
    else:
        form = EnregistrerFormulaireVenteForm()
    return render(request, 'enregistrer_formulaire_vente.html', {'form': form})

def actualiser_formulaire_vente(request, id_formulaire_vente):
    formulaire_vente = FormulaireVente.objects.get(pk=id_formulaire_vente)
    form = EnregistrerFormulaireVenteForm(request.POST or None, request.FILES or None,  instance=formulaire_vente)
    if form.is_valid():
        form.save()
        messages.success(request, "Le formulaire vente a été actualisé correctement.")
        return redirect('liste_des_formulaire_ventes')
    context = {'formulaire_vente': formulaire_vente, 'form': form}
    return render(request, 'store_app/actualizer_formulaire_vente.html', context)

def eliminer_formulaire_vente(request, id_formulaire_vente):
    formulaire_vente = get_object_or_404(FormulaireVente, id=id_formulaire_vente)
    formulaire_vente.delete()
    messages.success(request, "Le formulaire vente a été eliminer correctement.")
    return redirect('liste_des_formulaire_ventes')  # Reemplaza 'nombre_de_tu_vista' con el nombre de tu vista principal

def get_article_data(request, pk):
    try:
        article = FormulaireArticle.objects.get(pk=pk)
        data = {
            'description': article.description,
            'category': article.category.name,
            'prix': article.prix,
            'cout_revient_total': article.cout_revient_total
        }
        print( data)
        return JsonResponse(data)
    except FormulaireArticle.DoesNotExist:
        return JsonResponse({'error': 'Article not found'}, status=404)

def get_last_id(request):
    last_instance = FormulaireVente.objects.all().order_by('id').last()
    last_id = last_instance.id if last_instance else 0
    return JsonResponse({'last_id': last_id})
